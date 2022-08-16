from matplotlib import pyplot as plt
import pandas as pd
import numpy as np
import ipywidgets as widgets
from ipywidgets import interact, interactive, fixed, interact_manual
from ipywidgets import HBox, VBox
from IPython.display import display
from ipywidgets import Output, Tab

from models_class import DataParam
from models_class import Modelparam
from backtest import BackTest,MidBackTest

from random import seed
from random import randint

import plotly.express as px
import plotly.graph_objects as go

from openpyxl import load_workbook
wb = load_workbook(filename = 'DataDownload.xlsx')

'Data'
CURRENCY = wb.sheetnames[1:7]
from random import seed
from random import randint

###########################################################################################################################################
'widgets'
dropdown_curr = widgets.Dropdown(options=CURRENCY, description='Currency:',layout = widgets.Layout(width='200px'))

###########################################################################################################################################
'bottons'
"Add/Remove Model"
index = 0
mode_dic={}
model_storage = []

def add_model():
    out = Output()
    with out:
        m = Modelparam(randint(0,100).__str__())
        model_storage.append(m)
       
    return out

model_box = widgets.VBox([add_model()])
btn_add_model = widgets.Button(description = 'Add Model') 

def remove_model():
    rem = Output()
    with rem:
        remove = model_box.children[-1]
        model_box.children = model_box.children[:-1]
        remove.close()
        model_storage.pop()
        mode_dic.popitem()
    return rem
btn_rem_model = widgets.Button(description = 'Remove Model') 

"Save"
button_save = widgets.Button(
    description='OK',
    disabled=False,
    button_style='', # 'success', 'info', 'warning', 'danger' or ''
    tooltip='OK',
    layout = widgets.Layout(width='300px'),
    icon='' 
)

"Add/Remove Data"
data_storage = []
def add():
    out = Output()
    with out:
        m = DataParam()
        data_storage.append(m)
        
    return out

data_box = widgets.VBox([add()])
btn_add = widgets.Button(description = 'Add Scenario') 

def remove():
    rem = Output()
    with rem:
        remove = data_box.children[-1]
        data_box.children = data_box.children[:-1]
        remove.close()
        data_storage.pop()
        
    return rem
btn_rem = widgets.Button(description = 'Remove Scenario') 


    

"Run"
button_run1 = widgets.Button(
    description='Run',
    disabled=False,
    button_style='', # 'success', 'info', 'warning', 'danger' or ''
    tooltip='Run',
    layout = widgets.Layout(width='300px'),
    icon='' 
)



'Dashboard arrangement'
Add_remove_model = HBox([btn_add_model, btn_rem_model, button_save])
Add_remove_data = HBox([btn_add, btn_rem])
dashboard = widgets.VBox([dropdown_curr, model_box, Add_remove_model, data_box, Add_remove_data, button_run1])

    

###########################################################################################################################################
'Button Interactions'

def botton_addm_on_click(a):
    model_box.children=tuple(list(model_box.children) + [add_model()])
    return
btn_add_model.on_click(botton_addm_on_click)
index=index+1

def botton_remm_on_click(a):
    remove_model()
    return
btn_rem_model.on_click(botton_remm_on_click)

def botton_save_on_click(b):
    for model in model_storage:
        if model is None:
            print("model is none")
        else:
            model.Update()
            mode_dic[model.name]=model

            
               
button_save.on_click(botton_save_on_click)


def botton_run1_on_click(b):
    residual_error_dic={}
    hedgeratio_dic={}
    timestamp_dic={}
    timestamp=[]
    error_dic={}
    hedge_box=[]
    for data in data_storage:
        data.Update(mode_dic)
        name=data.target.__str__()+'y vs '+format(data.hedge)+ '[MOD'+data.i+" "+data.mode+']'
        if data.hedge not in hedge_box:
            hedge_box.append(data.hedge)
             
        if data.mode=='Selected Window':
            window=coeffpartgraph(dropdown_curr.value, data)
            residual_error_dic[name] = np.squeeze(window[1])
            error_dic[name]=window[2]
            timestamp=window[3]
            
        else:
            backtest=BackTest(dropdown_curr.value, data, mode_dic[data.index.value])
            residual_error_dic[name] = backtest['yvalues']
            hedgeratio_dic[name] = backtest['weighted_beta']
            timestamp = backtest['timestamp']
            error_dic[name]=backtest['errors']

    matrix_box = MidBackTest(dropdown_curr.value, hedge_box)

    plotTable(residual_error_dic, error_dic)

    plotResidualError(residual_error_dic,timestamp)

    plotHedgeRatio(hedgeratio_dic, timestamp)

    plotMatrix(matrix_box)
    
    

        
button_run1.on_click(botton_run1_on_click)


    
def botton_add_on_click(a):
    data_box.children=tuple(list(data_box.children) + [add()])
    return
btn_add.on_click(botton_add_on_click)

def botton_rem_on_click(a):
    remove()
    return
btn_rem.on_click(botton_rem_on_click)



#########################################################################################################
"Graphing"

"Plot Summary Table"
def plotTable(residual_error_dic, error_dic):
    dfff=plotSummary(residual_error_dic, error_dic)

    fig = go.Figure(data=[go.Table(
        header=dict(values=list(dfff.columns),
                    fill_color='#00ACFC',
                    align='center',
                font=dict(color='white')),
        cells=dict(values=[dfff['Name'], dfff['Mean'], dfff['Range'], dfff['Std'], dfff["Max Dropdown (weekly)"],dfff["Max Dropdown (monthly)"],dfff["0~0.05"],dfff["0.05~0.1"],dfff["0.1~0.15"],dfff["0.15~"]],
                fill_color='lavender',
                align='center'))
    ])
    
    fig.show()
    
"Plot Summary in Dataframe"    
import statistics
def plotSummary(residual_error_dic, error_dic):
    names = residual_error_dic.keys()
    SUMMARY = pd.DataFrame(columns=["Name", "Mean", "Range", "Std","Max Dropdown (weekly)","Max Dropdown (monthly)","0~0.05", "0.05~0.1", "0.1~0.15", "0.15~"])
    for name in names:
        df=np.array(error_dic[name])
        df_cum=np.array(residual_error_dic[name])

        std=round(df.std(),2)
        mean=round(df.mean(),2)
        max=round(df.max(),2)
        min=round(df.min(),2)
        r = '['+min.__str__() + ' , ' + max.__str__() + ']'
        values = df
        count1=0
        count2=0
        count3=0
        count4=0
        for val in values:
            if abs(val) > 0.05 and abs(val) <= 0.1:
                count2=count2+1
                
            if val <= 0.05 and val >= -0.05:
                count1=count1+1
            if abs(val) > 0.1 and abs(val) <= 0.15:
                count3=count3+1
            if abs(val) > 0.15:
                count4=count4+1

        P_range1=round(count1/len(values)*100,2)
        P1=P_range1.__str__()+'%'
        P_range2=round(count2/len(values)*100,2)
        P2=P_range2.__str__()+'%'
        P_range3=round(count3/len(values)*100,2)
        P3=P_range3.__str__()+'%'
        P_range4=round(count4/len(values)*100,2)
        P4=P_range4.__str__()+'%'
        
        
        
        '''max dropdown'''
        min_pnl_weekly=0
        for i in range(0, df_cum.shape[0]-10, 1):
            pnl = sum(df[i:i+10])
            if abs(pnl) > min_pnl_weekly:
                min_pnl_weekly = pnl.round(2)
                
        min_pnl_monthly=0
        for i in range(0, df_cum.shape[0]-44, 1):
            pnl = sum(df[i:i+44])
            if abs(pnl) > min_pnl_monthly:
                min_pnl_monthly = pnl.round(2)
        

        dict = {'Name': name, 'Mean': mean, 'Range': r,'Std': std, "Max Dropdown (weekly)":min_pnl_weekly,"Max Dropdown (monthly)":min_pnl_monthly,"0~0.05": P1, "0.05~0.1": P2, "0.1~0.15": P3, "0.15~": P4}
        SUMMARY = pd.concat([SUMMARY, pd.DataFrame.from_records([dict])])

    return SUMMARY


"plot Residual Errors on the same graph"
def plotResidualError(residual_error_dic, timestamp):
    df = pd.DataFrame(residual_error_dic)

    df.index= timestamp
    graphs=[]
    
    for scenario in df.columns:
        graphs.append(df.loc[:,scenario])
                
    fig = px.line(df, title="Residual Error")
    fig.update_xaxes(
                title_text = "Date",
                title_font = {"size": 15},
                title_standoff = 25)

    fig.update_yaxes(
            title_text = "Residual Error Index",
            title_font = {"size": 15},
            title_standoff = 25)
    
    figw = go.FigureWidget(fig)
    fig.show()
    
"Plot a matrix for 1-50 yrs"
def plotMatrix(matrix_box):
    fig_box=[]
    for i in matrix_box:
        pd_hedge=matrix_box[i].round(2)
        fig = go.Figure(data=[go.Table(
            header=dict(values=list(pd_hedge.columns),
                        fill_color='#00ACFC',
                        align='left'),
            cells=dict(values=[pd_hedge[col] for col in pd_hedge.columns],
                       fill_color='lavender',
                       align='left'))
        ], layout=layout)
        fig.show()

"plot hedge ratio over time for each scenarios"
def plotHedgeRatio(hedgeratio_dic, timestamp):
    keys = hedgeratio_dic.keys()
    for key in keys:
        df=hedgeratio_dic[key]
        df.index=timestamp
        
        fig = px.line(df, title="Hedge Ratio of "+ key)
        fig.update_xaxes(
                title_text = "Date",
                title_font = {"size": 15},
                title_standoff = 25)

        fig.update_yaxes(
                title_text = "Hedge Ratio Index",
                title_font = {"size": 15},
                title_standoff = 25)
        fig.show()
        
"plot selected window residual errors over time"
def coeffpartgraph(curr, data):
    raw_data = wb[curr]
    target, hedge_pt, start, end, window=data.target, data.hedge, data.s, data.e, data.window_size
    
    raw_data= pd.DataFrame(raw_data.values)
    raw_data.columns=raw_data.iloc[0,:]
    raw_data.index=raw_data.iloc[:,0]
    raw_data=raw_data.iloc[1:,1:20]
    raw_data=raw_data.replace('#VALUE!', np.NaN)
    raw_data=raw_data.dropna()

    X=raw_data.copy()
    X=X.diff()*100
    
    
    
    '''Target and Hedge points Settings'''
    X.index=pd.to_datetime(X.index)
    X=X.between_time('4:00:00','18:00:00')
    x=X.copy()
    x=x.loc['26/10/2020  3:00:00 AM':]
    X=X.loc[start+'  12:00:00 PM':end+'  5:00:00 PM']
    
    X['Date'] = X.index.date
    X=X.groupby('Date').sum()
    
    x['Date'] = x.index.date
    x=x.groupby('Date').sum()
    
    pd_hedge=X.loc[:,hedge_pt]
    np_hedge=np.asarray(pd_hedge)
    pd_target=X.loc[:,[target]]
    np_target=np.asarray(pd_target)
    
    reg=LinearRegression().fit(np_hedge,np_target)
    coeff=reg.coef_
    sum_coeff=coeff.sum()
    coeff_norm=coeff/sum_coeff
    
    np_weight=coeff_norm.transpose()
    
    pd_tempmtx=x.loc[:,hedge_pt]
    pd_tempcol=x.loc[:,[target]]

    np_tempmtx=np.asarray(pd_tempmtx)
    np_tempcol=np.asarray(pd_tempcol)

    model_Y=np.matmul(np_tempmtx,np_weight)
    model_error=np_tempcol-model_Y
    model_error_cumsum=np.cumsum(model_error,axis=0)
    yvalues=model_error_cumsum[window+1:]
    errors=model_error[window+1:]
    time_stamp_box=[]
    for i in x.index:
        time_stamp_box.append(i.strftime("%m/%d/%Y"))
    
    return coeff_norm, yvalues, errors, time_stamp_box[window+1:]
    
          
layout = go.Layout(
    autosize=False,
    width=400,
    height=600
)  

 

"Create a name for each scenario"
def format(hedge):
    str=''
    for i in range(len(hedge)):
        if i==len(hedge)-1:
            str=str+hedge[i].__str__()+'y '
        else:     
            str=str+hedge[i].__str__()+'y,'
    return str



display(dashboard)